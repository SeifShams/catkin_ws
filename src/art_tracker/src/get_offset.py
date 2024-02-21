#!/usr/bin/env python3
import rospy
import tf2_ros
from geometry_msgs.msg import TransformStamped
from openpyxl import Workbook

def main():
    rospy.init_node('tf_comparison_to_excel')

    # Create a TF2 Buffer and Listener
    tfBuffer = tf2_ros.Buffer()
    tfListener = tf2_ros.TransformListener(tfBuffer)

    # Define the names of the objects and the reference frame
    arUco1_frame = "ArUco1"
    arUco2_frame = "ArUco2"
    bod1_frame = "bod_2"
    bod2_frame = "bod_3"
    reference_frame = "world"

    # Create an Excel workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active

    # Add headers to the Excel sheet
    worksheet.cell(row=1, column=1, value="Object")
    worksheet.cell(row=1, column=2, value="X (m)")
    worksheet.cell(row=1, column=3, value="Y (m)")
    worksheet.cell(row=1, column=4, value="Z (m)")

    worksheet.cell(row=1, column=6, value="Object")
    worksheet.cell(row=1, column=7, value="X (m)")
    worksheet.cell(row=1, column=8, value="Y (m)")
    worksheet.cell(row=1, column=9, value="Z (m)")

    worksheet.cell(row=1, column=11, value="Object")
    worksheet.cell(row=1, column=12, value="X (m)")
    worksheet.cell(row=1, column=13, value="Y (m)")
    worksheet.cell(row=1, column=14, value="Z (m)")

    worksheet.cell(row=1, column=16, value="Object")
    worksheet.cell(row=1, column=17, value="X (m)")
    worksheet.cell(row=1, column=18, value="Y (m)")
    worksheet.cell(row=1, column=19, value="Z (m)")

    count = 0
    rate = rospy.Rate(1) # 1 Hz

    while not rospy.is_shutdown() and count != 10:
        
        count = count + 1
        try:
            # Get the transformations of object1 and object2 with respect to the reference frame
            tf_bod1 = tfBuffer.lookup_transform(reference_frame, bod1_frame, rospy.Time(0), rospy.Duration(1.0))
            tf_bod2 = tfBuffer.lookup_transform(reference_frame, bod2_frame, rospy.Time(0), rospy.Duration(1.0))

            tf_arUco1 = tfBuffer.lookup_transform(reference_frame, arUco1_frame, rospy.Time(0), rospy.Duration(1.0))
            tf_arUco2 = tfBuffer.lookup_transform(reference_frame, arUco2_frame, rospy.Time(0), rospy.Duration(1.0))

            # Extract translation data from the transformations
            x1 = tf_bod1.transform.translation.x
            y1 = tf_bod1.transform.translation.y
            z1 = tf_bod1.transform.translation.z

            x2 = tf_bod2.transform.translation.x
            y2 = tf_bod2.transform.translation.y
            z2 = tf_bod2.transform.translation.z

            x3 = tf_arUco1.transform.translation.x
            y3 = tf_arUco1.transform.translation.y
            z3 = tf_arUco1.transform.translation.z

            x4 = tf_arUco2.transform.translation.x
            y4 = tf_arUco2.transform.translation.y
            z4 = tf_arUco2.transform.translation.z

            # Write the data to the Excel sheet
            worksheet.append([  "Body 1", x1, y1, z1, "",
                                "Body 2", x2, y2, z2, "",
                                "ArUco 1", x3, y3, z3, "",
                                "ArUco 2", x4, y4, z4, "",])


            rospy.loginfo("Data " + str(count) + " recorded")

        except tf2_ros.LookupException as e:
            rospy.logerr(f"Error: {e}")

        rate.sleep() # Sleeps for 1/rate sec
    

    # Save the Excel file
    workbook.save("/home/danielvidal/catkin_ws/src/art_tracker/files/tf_comparison.xlsx")
    rospy.loginfo("TF comparison results written to tf_comparison.xlsx")

if __name__ == '__main__':
    main()
